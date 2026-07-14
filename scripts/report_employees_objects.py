"""Ежедневный (19:00 МСК) отчёт: лист «По сотрудникам» — грид-таблица
(сотрудник × объект за сегодня, плюс итоги за сегодня/месяц/всё время) и лист
«По объектам» (часы всего, кто работал). Собирается как .xlsx, заливается на
Bitrix24 Диск, в чат уходит ссылка на файл."""
from datetime import datetime, timedelta, timezone

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from bitrix_common import (
    TRIP_ENUM_ID, USER_NAMES, fmt_num, get_all_items, record_date,
    send_message, upload_to_disk, vacationing_users_on,
)

THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
TOTAL_FILL = PatternFill("solid", fgColor="F2F2F2")
BOLD = Font(bold=True)


def main():
    now_msk = datetime.now(timezone.utc) + timedelta(hours=3)
    today = now_msk.date()
    today_str = today.strftime("%d.%m.%Y")

    all_items = get_all_items()
    vac_today = vacationing_users_on(today, all_items)

    # Рабочие записи (без отпуска) — для часов и итогов.
    items = [it for it in all_items if it.get("ufCrm37Object") != "Отпуск"]
    for it in items:
        it["_date"] = record_date(it.get("title", ""))

    wb = openpyxl.Workbook()

    # === Лист 1: По сотрудникам (грид) ===
    ws1 = wb.active
    ws1.title = "По сотрудникам"

    # Матрица "сегодня": сотрудник -> объект -> {hours, trip}
    today_items = [
        it for it in items
        if it["_date"] == today
        and float(it.get("ufCrm37Hours") or 0) > 0
        and it.get("ufCrm37Object")
    ]
    matrix: dict = {}
    obj_set: set = set()
    for it in today_items:
        emp = it["assignedById"]
        obj = it["ufCrm37Object"]
        obj_set.add(obj)
        cell = matrix.setdefault(emp, {}).setdefault(obj, {"hours": 0.0, "trip": False})
        cell["hours"] += float(it["ufCrm37Hours"])
        if str(it.get("ufCrm37Place")) == str(TRIP_ENUM_ID):
            cell["trip"] = True

    special_order = ["Больничный", "Отгул"]

    def obj_key(o):
        return (special_order.index(o) if o in special_order else 99, o)

    objects = sorted(obj_set, key=obj_key)

    # Итоги по сотруднику: сегодня / за месяц / за всё время (по всей истории).
    totals_today: dict = {}
    totals_month: dict = {}
    totals_all: dict = {}
    emp_set: set = set()
    for it in items:
        d = it["_date"]
        if not d:
            continue
        emp = it["assignedById"]
        emp_set.add(emp)
        h = float(it.get("ufCrm37Hours") or 0)
        totals_all[emp] = totals_all.get(emp, 0.0) + h
        if d.year == today.year and d.month == today.month:
            totals_month[emp] = totals_month.get(emp, 0.0) + h
        if d == today:
            totals_today[emp] = totals_today.get(emp, 0.0) + h

    emp_set |= vac_today
    emps = sorted(emp_set, key=lambda e: USER_NAMES.get(e, str(e)))

    headers1 = ["№", "ФИО сотрудника"] + objects + ["Всего сегодня", "Всего за месяц", "Всего за всё время"]

    if not emps:
        ws1.cell(row=1, column=1, value="Записей пока нет.").font = BOLD
    else:
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers1))
        ws1.cell(row=1, column=1, value=f"Отчёт по сотрудникам — {today_str}").font = Font(bold=True, size=13)

        for col, h in enumerate(headers1, 1):
            c = ws1.cell(row=2, column=col, value=h)
            c.font = BOLD
            c.fill = HEADER_FILL
            c.border = BORDER
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        col_totals = {o: 0.0 for o in objects}
        grand_today = 0.0
        grand_month = 0.0
        grand_all = 0.0
        row = 3
        for i, emp in enumerate(emps, 1):
            name = USER_NAMES.get(emp, f"ID{emp}")
            ws1.cell(row=row, column=1, value=i)
            ws1.cell(row=row, column=2, value=name)

            on_vacation_now = emp in vac_today and emp not in matrix
            if on_vacation_now:
                for ci in range(len(objects)):
                    ws1.cell(row=row, column=3 + ci, value="")
                ws1.cell(row=row, column=3 + len(objects), value="Отпуск")
            else:
                for ci, o in enumerate(objects):
                    cell = matrix.get(emp, {}).get(o)
                    if cell:
                        val = fmt_num(cell["hours"]) + (" (к)" if cell["trip"] else "")
                        col_totals[o] += cell["hours"]
                        ws1.cell(row=row, column=3 + ci, value=val)
                    else:
                        ws1.cell(row=row, column=3 + ci, value="")
                ws1.cell(row=row, column=3 + len(objects), value=fmt_num(totals_today.get(emp, 0.0)))

            month_h = totals_month.get(emp, 0.0)
            all_h = totals_all.get(emp, 0.0)
            ws1.cell(row=row, column=4 + len(objects), value=fmt_num(month_h))
            ws1.cell(row=row, column=5 + len(objects), value=fmt_num(all_h))

            grand_today += totals_today.get(emp, 0.0)
            grand_month += month_h
            grand_all += all_h

            for col in range(1, len(headers1) + 1):
                c = ws1.cell(row=row, column=col)
                c.border = BORDER
                c.alignment = Alignment(horizontal="center", vertical="center")
            ws1.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="center")
            row += 1

        ws1.cell(row=row, column=2, value="Всего на объект / ИТОГО")
        for ci, o in enumerate(objects):
            ws1.cell(row=row, column=3 + ci, value=fmt_num(col_totals[o]))
        ws1.cell(row=row, column=3 + len(objects), value=fmt_num(grand_today))
        ws1.cell(row=row, column=4 + len(objects), value=fmt_num(grand_month))
        ws1.cell(row=row, column=5 + len(objects), value=fmt_num(grand_all))
        for col in range(1, len(headers1) + 1):
            c = ws1.cell(row=row, column=col)
            c.border = BORDER
            c.font = BOLD
            c.fill = TOTAL_FILL
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws1.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="center")

        ws1.column_dimensions["A"].width = 5
        ws1.column_dimensions["B"].width = 24
        for i in range(len(objects)):
            ws1.column_dimensions[get_column_letter(3 + i)].width = 16
        ws1.column_dimensions[get_column_letter(3 + len(objects))].width = 13
        ws1.column_dimensions[get_column_letter(4 + len(objects))].width = 13
        ws1.column_dimensions[get_column_letter(5 + len(objects))].width = 16

    # === Лист 2: По объектам (без изменений) ===
    ws2 = wb.create_sheet("По объектам")
    ws2.cell(row=1, column=1, value=f"Отчёт по объектам — всего (на {today_str})").font = Font(bold=True, size=13)
    headers2 = ["Объект", "Часы всего", "Сотрудники"]
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(row=2, column=col, value=h)
        c.font = BOLD
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")

    by_obj: dict = {}
    for it in items:
        obj = it.get("ufCrm37Object")
        if not obj:
            continue
        h = float(it.get("ufCrm37Hours") or 0)
        d = by_obj.setdefault(obj, {"hours": 0.0, "emps": set()})
        d["hours"] += h
        d["emps"].add(USER_NAMES.get(it["assignedById"], f"ID{it['assignedById']}"))

    row = 3
    grand = 0.0
    for obj, d in sorted(by_obj.items(), key=lambda kv: -kv[1]["hours"]):
        ws2.cell(row=row, column=1, value=obj)
        ws2.cell(row=row, column=2, value=fmt_num(d["hours"]))
        ws2.cell(row=row, column=3, value=", ".join(sorted(d["emps"])))
        for col in range(1, 4):
            ws2.cell(row=row, column=col).border = BORDER
        grand += d["hours"]
        row += 1

    ws2.cell(row=row, column=1, value="ИТОГО").font = BOLD
    ws2.cell(row=row, column=2, value=fmt_num(grand)).font = BOLD
    for col in range(1, 4):
        ws2.cell(row=row, column=col).border = BORDER
        ws2.cell(row=row, column=col).fill = TOTAL_FILL

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 40

    filename = f"Отчет_по_сотрудникам_и_объектам_{today_str}.xlsx"
    wb.save(filename)
    with open(filename, "rb") as f:
        content = f.read()
    url = upload_to_disk(filename, content)
    send_message("chat66269", f"[B]📊 Отчёт по сотрудникам и объектам — {today_str}[/B]\n{url}")
    print(f"OK: {url}")


if __name__ == "__main__":
    main()
