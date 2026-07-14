"""Ежедневный (20:00 МСК) матричный отчёт «сотрудник × объект» за сегодня.
Собирается как настоящий .xlsx, заливается на Bitrix24 Диск, в чат уходит
только ссылка на файл (не текст)."""
from datetime import datetime, timedelta, timezone

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from bitrix_common import (
    TRIP_ENUM_ID, USER_NAMES, fmt_num, get_all_items, record_date,
    send_message, upload_to_disk, vacationing_users_on,
)

THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
TOTAL_FILL = PatternFill("solid", fgColor="F2F2F2")
BOLD = Font(bold=True)


def main():
    now_msk = datetime.now(timezone.utc) + timedelta(hours=3)
    today = now_msk.date()
    today_str = today.strftime("%d.%m.%Y")

    items = get_all_items()
    vac_today = vacationing_users_on(today, items)

    today_items = [
        it for it in items
        if record_date(it.get("title", "")) == today
        and float(it.get("ufCrm37Hours") or 0) > 0
        and it.get("ufCrm37Object")
        and it["ufCrm37Object"] != "Отпуск"
    ]

    matrix: dict = {}
    obj_set: set = set()
    emp_set: set = set()
    for it in today_items:
        emp = it["assignedById"]
        obj = it["ufCrm37Object"]
        emp_set.add(emp)
        obj_set.add(obj)
        cell = matrix.setdefault(emp, {}).setdefault(obj, {"hours": 0.0, "trip": False})
        cell["hours"] += float(it["ufCrm37Hours"])
        if str(it.get("ufCrm37Place")) == str(TRIP_ENUM_ID):
            cell["trip"] = True

    emp_set |= vac_today

    special_order = ["Больничный", "Отгул"]

    def obj_key(o):
        return (special_order.index(o) if o in special_order else 99, o)

    objects = sorted(obj_set, key=obj_key)
    emps = sorted(emp_set, key=lambda e: USER_NAMES.get(e, str(e)))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Матрица"

    headers = ["№", "ФИО сотрудника"] + objects + ["Всего"]

    if not emps:
        ws.cell(row=1, column=1, value="Сегодня никто ещё не заполнил часы.").font = BOLD
    else:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.cell(row=1, column=1, value=f"Отчёт за {today_str}").font = Font(bold=True, size=13)

        for col, h in enumerate(headers, 1):
            c = ws.cell(row=2, column=col, value=h)
            c.font = BOLD
            c.fill = HEADER_FILL
            c.border = BORDER
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        col_totals = {o: 0.0 for o in objects}
        row_idx = 3
        for i, emp in enumerate(emps, 1):
            name = USER_NAMES.get(emp, f"ID{emp}")
            ws.cell(row=row_idx, column=1, value=i)
            ws.cell(row=row_idx, column=2, value=name)
            if emp in vac_today and emp not in matrix:
                for ci in range(len(objects)):
                    ws.cell(row=row_idx, column=3 + ci, value="")
                ws.cell(row=row_idx, column=len(headers), value="Отпуск")
            else:
                row_total = 0.0
                for ci, o in enumerate(objects):
                    cell = matrix.get(emp, {}).get(o)
                    if cell:
                        val = fmt_num(cell["hours"]) + (" (к)" if cell["trip"] else "")
                        row_total += cell["hours"]
                        col_totals[o] += cell["hours"]
                        ws.cell(row=row_idx, column=3 + ci, value=val)
                    else:
                        ws.cell(row=row_idx, column=3 + ci, value="")
                ws.cell(row=row_idx, column=len(headers), value=fmt_num(row_total))
            for col in range(1, len(headers) + 1):
                c = ws.cell(row=row_idx, column=col)
                c.border = BORDER
                c.alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="left", vertical="center")
            row_idx += 1

        ws.cell(row=row_idx, column=2, value="Всего на объект")
        for ci, o in enumerate(objects):
            ws.cell(row=row_idx, column=3 + ci, value=fmt_num(col_totals[o]))
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=row_idx, column=col)
            c.border = BORDER
            c.font = BOLD
            c.fill = TOTAL_FILL
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="left", vertical="center")

        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 24
        for i in range(len(objects)):
            ws.column_dimensions[get_column_letter(3 + i)].width = 16
        ws.column_dimensions[get_column_letter(3 + len(objects))].width = 12

    filename = f"Матрица_{today_str}.xlsx"
    wb.save(filename)
    with open(filename, "rb") as f:
        content = f.read()
    url = upload_to_disk(filename, content)
    send_message("chat66305", f"[B]📅 Отчёт за {today_str}[/B]\n{url}")
    print(f"OK: {url}")


if __name__ == "__main__":
    main()
